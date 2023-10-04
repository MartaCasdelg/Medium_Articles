# Import libraries
import psycopg2
import openpyxl

# Create a connection to existing DB
connection = psycopg2.connect(
    database = 'HR_Data',
    user = 'postgres',
    password = '**********',
    host = 'localhost',
    port = '5432'
)

# Open a cursor object to perform database operations
cursor = connection.cursor()

# Save Excel file location into a variable
excel_file = 'C:\\Users\\Documents\\HR_Employees\\HR_Employee_Data.xlsx'

# Open the Excel workbook and load the active sheet into a variable
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Create a list with the column names in the first row of the excel workbook
column_names = [column.value for column in sheet[1]]

# Create an empty list 
data = []
# Create a loop to iterate over all rows in the excel sheet (except the titles) and append the data to the list
for row in sheet.iter_rows(min_row = 2, values_only = True): # iter_rows is an openpyxl function
    data.append(row)

# Set a name for the postgreSQL schema and table where we will put the data
schema_name = 'hr_employees'
table_name = 'evaluation'

# Write a query to create a schema using schema_name. Save it into a variable.
schema_creation_query = f'CREATE SCHEMA IF NOT EXISTS {schema_name}'

# Write a query to create a table (table_name) in the schema (schema_name). It must contain all columns in column_names. Save it into a variable.
table_creation_query = f"""
    CREATE TABLE IF NOT EXISTS {schema_name}.{table_name} (
        {", ".join([f'"{name}" TEXT' for name in column_names])}
    )
"""

# Use the cursor to execute both queries
cursor.execute(schema_creation_query)
cursor.execute(table_creation_query)

# Create a parameterized SQL query to insert the data into the table
insert_data_query = f"""
    INSERT INTO {schema_name}.{table_name} ({", ".join([f'"{name}"' for name in column_names])})
    VALUES ({", ".join(['%s' for _ in column_names])})
"""

# Execute the query using the data list as parameter
cursor.executemany(insert_data_query, data)

# Make the changes to the database persistent
connection.commit()

# Close communication with the database
cursor.close()
connection.close()

# Print a message
print('Import successfully completed!')